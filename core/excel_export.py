from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING

import pandas as pd
from PyQt5.QtWidgets import QFileDialog, QMessageBox

if TYPE_CHECKING:
    from core.widget import OrderDashboardWidget


def export_excel(widget: OrderDashboardWidget):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side

    vendor_buttons = {
        "코스트코": widget.on_click_filter_costco,
        "이마트": widget.on_click_filter_emart,
        "홈플/컬리": widget.on_click_filter_hk,
        "롯데": widget.on_click_filter_lotte,
    }
    vendors = list(vendor_buttons.keys())

    today_str = datetime.now().strftime("%Y%m%d_%H%M")
    default_name = f"제품현황_업체별_{today_str}.xlsx"

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    default_path = os.path.join(desktop, default_name)

    path, _ = QFileDialog.getSaveFileName(
        widget, "업체별 제품현황 엑셀 저장", default_path,
        "Excel Files (*.xlsx);;All Files (*)",
    )
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:

            for vendor in vendors:
                vendor_buttons[vendor]()
                table = widget.ui.tableWidget1

                if table.rowCount() == 0:
                    continue

                headers = []
                for c in range(table.columnCount()):
                    item = table.horizontalHeaderItem(c)
                    headers.append(item.text() if item else f"열{c + 1}")

                rows = []
                for r in range(table.rowCount()):
                    row_vals = []
                    for c in range(table.columnCount()):
                        item = table.item(r, c)
                        row_vals.append(item.text() if item else "")
                    rows.append(row_vals)

                df = pd.DataFrame(rows, columns=headers)
                safe_name = vendor.replace("/", "_")
                df.to_excel(writer, sheet_name=safe_name, index=False)

                wb = writer.book
                ws = wb[safe_name]

                header_font = Font(bold=True)
                header_align = Alignment(horizontal="center", vertical="center")
                left_align = Alignment(horizontal="left", vertical="center")
                right_align = Alignment(horizontal="right", vertical="center")
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for col_idx, col_name in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = border

                    max_len = len(str(col_name))
                    col_series = df[col_name].astype(str)
                    if not col_series.empty:
                        max_len = max(max_len, col_series.map(len).max())
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = border
                        if col_idx in (1, 2):
                            cell.alignment = left_align
                        else:
                            cell.alignment = right_align

        QMessageBox.information(widget, "완료", f"엑셀 파일이 저장되었습니다.\n{path}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        QMessageBox.critical(widget, "오류", f"엑셀 저장 중 오류 발생\n{e}")
