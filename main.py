import sys
import pandas as pd

from PyQt5.QtCore import Qt, QDate, QTimer
from math import ceil
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QTableWidgetItem,
    QMessageBox,
    QAbstractItemView,
    QTableWidget,
    QHeaderView,
    QDialog,
    QFileDialog,
    QInputDialog,
    QLineEdit,
)
from PyQt5.QtGui import QBrush, QColor, QFont

from datetime import datetime
from UTIL.db_handler import getdb, runquery, closedb, db_connection
from ci_cd.updatedown import check_version_and_update
from UTIL.util import fmt
from logic.cal_values import (
    calc_order_qty_packs,
    get_pacsu_by_co,
    get_produced_qty_packs,
    calc_plan_kg_by_recipe,
    get_stock_from_pan,
    get_prev_residue_from_today,
    recalc_dashboard_raw_keep_manual,
    recalc_dashboard_sauce_keep_manual,
    recalc_dashboard_vege_keep_manual,
    calc_trate_value,
)
from UTIL.const import (
    COL_VENDOR, COL_PRODUCT, COL_DEADLINE, COL_PKG,
    COL_ORDER, COL_FINAL_ORDER, COL_DIFF, COL_PREV_RES,
    COL_PRODUCTION, COL_PLAN, COL_PLAN_KG, COL_CUR_PROD,
    COL_SHIPMENT_TIME, COL_TODAY_RES, COL_TRATE, COL_WORK_STATUS,
    DB_NAME,
)

from UI.dashboard import Ui_Form
from dialog.DashboardLogDialog import DashboardLogDialog
from dialog.ProductListDialog import ProductListDialog
from dialog.ProductNameDialog import ProductNameDialog
from UTIL.db_product_handler import fetch_default_products

CURRENT_VERSION = "a-0031"
PROGRAM_NAME = "factory_dashboard"


# ---------------------------------------------------------
# 메인 위젯
# ---------------------------------------------------------
class OrderDashboardWidget(QWidget):

    # =====================================================
    # 1. 초기화 & 기본 기능
    # =====================================================
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # 상태 변수 (기존 global → 인스턴스)
        self.current_level = 0
        self.current_user = ""
        try:
            self.product_list = fetch_default_products()
        except Exception:
            self.product_list = []

        # 페이지네이션 상태
        self.product_page = 0
        self.product_page_size = 14
        self.product_total_pages = 1

        # 현재 선택된 업체 (기본값: 코스트코)
        self.current_vendor = "코스트코"
        self.ui.tab_frequency.setText("30")

        self._fullscreen_mode = False
        self.ui.control_frame.hide()
        self.show_hidden = False

        # 날짜 설정
        self.ui.dateEdit.setDate(QDate.currentDate())
        qdate = self.ui.dateEdit.date()
        date_str = qdate.toString("yyyy-MM-dd")
        weekday_str = qdate.toString("ddd")
        self.ui.dateText.setText(f"{date_str} ({weekday_str})")

        # 변경 이벤트 플래그
        self._item_changed_connected = {
            "product": False,
            "raw": False,
            "sauce": False,
            "vege": False,
        }

        # 테이블 스타일 적용
        for table in self._all_tables():
            self._setup_table_base(table)

        # 페이지네이션 바 생성 (tab1 > verticalLayout_3 > tableWidget1 아래)
        self._setup_pagination_bar()

        # 품명 매핑 캐시
        self.uname_map_cache = {}
        self.refresh_uname_map_cache()

        # 타이머 설정 (화면 전환/갱신용)
        self.is_auto_rotation = False
        self.vendors_rotation = ["코스트코", "이마트", "홈플/컬리", "롯데"]
        self.rotation_index = 0

        self.timer_view = QTimer(self)
        self.timer_view.timeout.connect(self._on_timer_tick)
        self.timer_view.start(1000 * 30)

        # 30분 자동 갱신 타이머
        self.timer_30min = QTimer(self)
        self.timer_30min.timeout.connect(self._auto_update_every_30min)
        self.timer_30min.start(1000 * 60 * 30)

        # 버튼 / 시그널 연결
        self.ui.btn_view.clicked.connect(self.on_click_toggle_fullscreen)
        self.ui.btn_prev.clicked.connect(self.on_click_prev_date)
        self.ui.btn_next.clicked.connect(self.on_click_next_date)

        self.ui.btn_product.clicked.connect(self.on_click_tab_product)
        self.ui.btn_raw.clicked.connect(self.on_click_tab_raw)
        self.ui.btn_sauce.clicked.connect(self.on_click_tab_sauce)
        self.ui.btn_vege.clicked.connect(self.on_click_tab_vege)

        # 업체 필터 버튼 연결
        self.ui.btn_costco.clicked.connect(self.on_click_filter_costco)
        self.ui.btn_emart.clicked.connect(self.on_click_filter_emart)
        self.ui.btn_homeplus.clicked.connect(self.on_click_filter_hk)
        self.ui.btn_lotte.clicked.connect(self.on_click_filter_lotte)

        self.ui.btn_add.clicked.connect(self.on_click_add_dummy_rows)
        self.ui.btn_del.clicked.connect(self.on_click_delete_rows)
        self.ui.btn_del_row.clicked.connect(self.on_click_delete_selected_products)
        self.ui.btn_update.clicked.connect(self.on_click_update_order_qty_after)
        self.ui.btn_update_product.clicked.connect(self.on_click_update_product)
        self.ui.btn_log.clicked.connect(self.on_click_show_log_dialog)
        self.ui.btn_excel.clicked.connect(self.on_click_export_excel)
        self.ui.btn_admin.clicked.connect(self.on_click_toggle_admin)
        self.ui.btn_complete.clicked.connect(self.on_click_complete_product)
        self.ui.btn_custom.clicked.connect(self.on_click_custom)
        self.ui.btn_renew.clicked.connect(self._renew_values_manually)

        self.ui.btn_hide_row.clicked.connect(self.on_click_hide_row)
        self.ui.btn_show_hide.clicked.connect(self.on_click_toggle_show_hide)
        self.ui.ml_check.stateChanged.connect(self._load_product_tab)

        # 화면 전환/고정 버튼
        self.ui.btn_autoPage.setText("화면고정")
        self.ui.btn_autoPage.clicked.connect(self.on_click_toggle_mode)

        # 탭 전환
        self.ui.tabWidget.currentChanged.connect(self.on_tab_changed)

        # 날짜 변경 이벤트
        self.ui.dateEdit.dateChanged.connect(self.on_date_changed)

        # 최초 로딩
        self._load_product_tab()
        self._apply_column_visibility_rules()

    # ---------------------------------------------------------
    # 유틸리티 헬퍼
    # ---------------------------------------------------------
    def _all_tables(self):
        return [self.ui.tableWidget1, self.ui.tableWidget2,
                self.ui.tableWidget3, self.ui.tableWidget4]

    def _material_table(self, tab_key: str) -> QTableWidget:
        return {"raw": self.ui.tableWidget2,
                "sauce": self.ui.tableWidget3,
                "vege": self.ui.tableWidget4}[tab_key]

    def _material_db_table(self, tab_key: str) -> str:
        return {"raw": "DASHBOARD_RAW",
                "sauce": "DASHBOARD_SAUCE",
                "vege": "DASHBOARD_VEGE"}[tab_key]

    def _material_row_height(self, tab_key: str) -> int:
        return {"raw": 50, "sauce": 46, "vege": 46}[tab_key]

    def _tab_loader(self, idx: int):
        loaders = {
            0: self._load_product_tab,
            1: lambda: self._load_material_tab("raw"),
            2: lambda: self._load_material_tab("sauce"),
            3: lambda: self._load_material_tab("vege"),
        }
        loader = loaders.get(idx)
        if loader:
            loader()

    def _reload_all_tabs(self):
        self._load_product_tab()
        self._load_material_tab("raw")
        self._load_material_tab("sauce")
        self._load_material_tab("vege")

    # =====================================================
    # 2. UI 상태 관련 함수
    # =====================================================
    def on_click_toggle_fullscreen(self):
        self._fullscreen_mode = not self._fullscreen_mode

        if self._fullscreen_mode:
            self.showFullScreen()
            self.ui.view_frame.hide()
            self.ui.view_frame2.hide()
            self.ui.control_frame.hide()
        else:
            self.showNormal()
            self.ui.view_frame.show()
            self.ui.view_frame2.show()
            if self.current_level >= 2:
                self.ui.control_frame.show()
            else:
                self.ui.control_frame.hide()

        self.layout().update()

    def _ask_admin_login(self):
        """DASHBOARD_ID 테이블에서 비밀번호만으로 사용자 검증."""
        pw, ok = QInputDialog.getText(
            self, "관리자 로그인", "비밀번호를 입력하세요:", QLineEdit.Password
        )
        if not ok:
            return False

        try:
            with db_connection("GP") as (conn, cur):
                sql = "SELECT name, level FROM DASHBOARD_ID WHERE pw = %s"
                df = runquery(cur, sql, [pw])
        except (ConnectionError, Exception) as e:
            QMessageBox.warning(self, "로그인 실패", f"DB 연결 실패: {e}")
            return False

        if df is None or df.empty:
            QMessageBox.warning(self, "로그인 실패", "일치하는 계정이 없습니다.")
            return False

        self.current_user = str(df.iloc[0]["name"]).strip()
        self.current_level = int(df.iloc[0]["level"])
        return True

    def on_click_toggle_admin(self):
        # 이미 관리자면 OFF
        if self.current_level >= 1:
            self.current_level = 0
            self.current_user = ""
            self.ui.control_frame.hide()
            self._apply_column_visibility_rules()
            self.ui.btn_admin.setText("관리자")
            self._reload_all_tabs()
            return

        # 로그인 시도
        if self._ask_admin_login():
            if self.current_level >= 2:
                self.ui.control_frame.show()
            else:
                self.ui.control_frame.hide()

            self._apply_column_visibility_rules()
            self.ui.btn_admin.setText(f"관리자: {self.current_user}")
            self._reload_all_tabs()

    def on_click_custom(self):
        """품명 관리 다이얼로그 오픈"""
        dlg = ProductNameDialog(self)
        dlg.exec_()

    def refresh_uname_map_cache(self):
        """Dashboard_UNAME_MAP 테이블에서 매핑 정보 로드하여 캐시 갱신"""
        self.uname_map_cache = {}
        try:
            with db_connection(DB_NAME) as (conn, cur):
                sql = "SELECT before_value, after_value FROM Dashboard_UNAME_MAP"
                df = runquery(cur, sql)
                if df is not None and not df.empty:
                    for _, row in df.iterrows():
                        bf = str(row['before_value']).strip()
                        af = row['after_value']
                        if af is not None:
                            af = str(af).strip()
                            if af:
                                self.uname_map_cache[bf] = af
        except Exception as e:
            print(f"매핑 캐시 로드 실패: {e}")

        if hasattr(self, 'ui'):
            self._load_product_tab()

    def logout_if_logged_in(self):
        if self.current_level >= 1:
            self.on_click_toggle_admin()

    def on_click_hide_row(self):
        """선택된 제품행의 hide 필드를 0/1 토글"""
        table = self.ui.tableWidget1
        selected_rows = sorted({idx.row() for idx in table.selectedIndexes()})

        if not selected_rows:
            QMessageBox.information(self, "안내", "숨김 처리할 제품을 선택하세요.")
            return

        with db_connection(DB_NAME) as (conn, cur):
            for row in selected_rows:
                item = table.item(row, 0)
                if not item:
                    continue

                pk = item.data(Qt.UserRole)
                if not pk:
                    continue

                df = runquery(cur, "SELECT hide FROM ORDER_DASHBOARD WHERE PK = %s", [pk])
                if df is None or df.empty:
                    continue

                cur_val = df.iloc[0]["hide"]
                new_val = 1 if cur_val is None else (0 if int(cur_val) == 1 else 1)
                runquery(cur, "UPDATE ORDER_DASHBOARD SET hide = %s WHERE PK = %s", [new_val, pk])

            QMessageBox.information(self, "완료", "선택한 제품의 hide 값이 변경되었습니다.")

        self._load_product_tab()

    def on_click_toggle_show_hide(self):
        self.show_hidden = not self.show_hidden
        self.ui.btn_show_hide.setText("숨김포함" if self.show_hidden else "숨김제외")
        self._load_product_tab()

    # =====================================================
    # 3. 탭 / 날짜 이동
    # =====================================================
    def on_click_prev_date(self):
        self.ui.dateEdit.setDate(self.ui.dateEdit.date().addDays(-1))

    def on_click_next_date(self):
        self.ui.dateEdit.setDate(self.ui.dateEdit.date().addDays(1))

    def on_date_changed(self):
        self.product_page = 0
        qdate = self.ui.dateEdit.date()
        date_str = qdate.toString("yyyy-MM-dd")
        weekday_str = qdate.toString("ddd")
        self.ui.dateText.setText(f"{date_str} ({weekday_str})")
        self._tab_loader(self.ui.tabWidget.currentIndex())

    def on_click_tab_product(self):
        self.ui.tabWidget.setCurrentIndex(0)

    def on_click_tab_raw(self):
        self.ui.tabWidget.setCurrentIndex(1)

    def on_click_tab_sauce(self):
        self.ui.tabWidget.setCurrentIndex(2)

    def on_click_tab_vege(self):
        self.ui.tabWidget.setCurrentIndex(3)

    def on_tab_changed(self, idx: int):
        self._tab_loader(idx)

    # =====================================================
    # 타이머 & 화면 전환 로직
    # =====================================================
    def _get_frequency(self) -> int:
        try:
            text = self.ui.tab_frequency.text().strip()
            val = int(text) if text else 30
        except ValueError:
            val = 30
        return max(val, 5)

    def _on_timer_tick(self):
        freq_sec = self._get_frequency()
        new_interval = freq_sec * 1000
        if self.timer_view.interval() != new_interval:
            self.timer_view.setInterval(new_interval)

        if self.is_auto_rotation:
            # 현재 업체에 다음 페이지가 있으면 페이지만 넘김
            if self.product_page < self.product_total_pages - 1:
                self.product_page += 1
                self._load_product_tab()
            else:
                # 마지막 페이지였으면 다음 업체로 이동 (페이지 리셋)
                self.rotation_index = (self.rotation_index + 1) % len(self.vendors_rotation)
                next_vendor = self.vendors_rotation[self.rotation_index]
                self.ui.label_retailer.setText(next_vendor)
                self._change_vendor_filter(next_vendor)
        else:
            idx = self.ui.tabWidget.currentIndex()
            if idx == 0:
                self._load_product_tab()

    def on_click_toggle_mode(self):
        """화면고정 <-> 자동전환 토글"""
        self.is_auto_rotation = not self.is_auto_rotation

        if self.is_auto_rotation:
            self.ui.btn_autoPage.setText("자동전환")
            try:
                self.rotation_index = self.vendors_rotation.index(self.current_vendor)
            except ValueError:
                self.rotation_index = 0
            self._on_timer_tick()
        else:
            self.ui.btn_autoPage.setText("화면고정")
            self._on_timer_tick()

    def _auto_update_every_30min(self):
        print(f"[_auto_update_every_30min] {datetime.now()} 자동 갱신 시작")
        self.on_click_update_order_qty_after(silent=True)
        self.on_click_update_product(silent=True)
        self.logout_if_logged_in()

    def _renew_values_manually(self):
        print(f"[_renew_values_manually] {datetime.now()} 수동 갱신 시작")
        self.on_click_update_order_qty_after(silent=True)
        self.on_click_update_product(silent=True)

    # =====================================================
    # 업체 필터링 (제품 탭)
    # =====================================================
    def _change_vendor_filter(self, vendor_name: str):
        self.current_vendor = vendor_name
        self.product_page = 0
        self._load_product_tab()

    def on_click_filter_costco(self):
        self._change_vendor_filter("코스트코")

    def on_click_filter_emart(self):
        self._change_vendor_filter("이마트")

    def on_click_filter_hk(self):
        self._change_vendor_filter("홈플/컬리")

    def on_click_filter_lotte(self):
        self._change_vendor_filter("롯데")

    # =====================================================
    # 4. 테이블 UI 설정
    # =====================================================
    def _setup_pagination_bar(self):
        """dashboard.ui의 pagination 프레임 초기화 및 시그널 연결"""
        self.ui.pagination.hide()
        self.ui.btn_left.clicked.connect(self._on_page_prev)
        self.ui.btn_right.clicked.connect(self._on_page_next)

    def _update_pagination_ui(self):
        """페이지 수에 따라 pagination 프레임 표시/숨김 및 버튼 상태 업데이트"""
        if self.product_total_pages <= 1:
            self.ui.pagination.hide()
            return
        self.ui.pagination.show()
        self.ui.page_label.setText(f"{self.product_page + 1} / {self.product_total_pages}")
        self.ui.btn_left.setEnabled(self.product_page > 0)
        self.ui.btn_right.setEnabled(self.product_page < self.product_total_pages - 1)

    def _on_page_prev(self):
        if self.product_page > 0:
            self.product_page -= 1
            self._load_product_tab()

    def _on_page_next(self):
        if self.product_page < self.product_total_pages - 1:
            self.product_page += 1
            self._load_product_tab()

    def _setup_table_base(self, table: QTableWidget):
        table.setAlternatingRowColors(True)
        table.setShowGrid(True)
        table.setEditTriggers(QAbstractItemView.DoubleClicked)
        table.setStyleSheet("""
            QTableWidget {
                font-size: 20px;
                alternate-background-color: #f6f7fb;
                gridline-color: #c0c0c0;
            }
            QHeaderView::section {
                font-size: 20px;
                font-weight: bold;
                color: black;
                padding: 5px;
                border: 1px solid #a0a0a0;
            }
        """)
        table.verticalHeader().setDefaultSectionSize(60)
        table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

    def _setup_product_headers(self, table):
        headers = [
            "업체명", "품명", "소비기한", "팩중량", "발주량",
            "최종발주", "팩 차이", "전일 잔피", "생산 팩수", "생산계획",
            "팩수 to kg", "데크출고", "최근출고", "당일 잔피", "수율", "작업상태",
        ]

        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        header_normal = QColor("#8fbcd4")
        header_live = QColor("#b7d9ff")

        for col in range(len(headers)):
            item = table.horizontalHeaderItem(col)
            if not item:
                continue
            if col in (COL_FINAL_ORDER, COL_CUR_PROD, COL_SHIPMENT_TIME):
                item.setBackground(QBrush(header_live))
            else:
                item.setBackground(QBrush(header_normal))

    def _setup_material_headers(self, table: QTableWidget):
        """Raw/Sauce/Vege 공통 헤더"""
        headers = [
            "품명", "재고량", "예상발주량", "최종발주량",
            "선 생산량", "예상부족량", "입고예정량", "예상재고",
        ]

        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        header_normal = QColor("#8fbcd4")
        header_edit = QColor("#ffdd99")

        for col in range(len(headers)):
            item = table.horizontalHeaderItem(col)
            if not item:
                continue
            if col in (4, 6):
                item.setBackground(QBrush(header_edit))
            else:
                item.setBackground(QBrush(header_normal))

    def _create_cell(
            self,
            text: str,
            pk: int,
            alignment: Qt.AlignmentFlag,
            *,
            editable: bool = False,
            underline: bool = False,
            foreground: QColor | None = None,
    ) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setData(Qt.UserRole, pk)

        font = QFont()
        font.setPointSize(20 if self.current_level >= 1 else 24)
        font.setUnderline(underline)
        item.setFont(font)

        item.setTextAlignment(alignment)

        base_flags = item.flags()
        if editable and self.current_level >= 1:
            item.setFlags(base_flags | Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            item.setForeground(QBrush(QColor("#777777")))
        else:
            item.setFlags(base_flags & ~Qt.ItemIsEditable)

        if foreground is not None:
            item.setForeground(QBrush(foreground))

        return item

    def _create_product_item(self, text: str, pk: int, col: int):
        if col in (COL_WORK_STATUS, COL_DEADLINE, COL_SHIPMENT_TIME):
            alignment = Qt.AlignCenter
        elif col in (COL_VENDOR, COL_PRODUCT):
            alignment = Qt.AlignLeft | Qt.AlignVCenter
        else:
            alignment = Qt.AlignRight | Qt.AlignVCenter

        editable = col in {COL_PLAN, COL_TODAY_RES, COL_PREV_RES}
        foreground = QColor("#0066cc") if col in (COL_FINAL_ORDER, COL_CUR_PROD) else None

        return self._create_cell(
            text=text, pk=pk, alignment=alignment,
            editable=editable, foreground=foreground,
        )

    def _create_material_item(self, text: str, pk: int, col: int):
        """Raw/Sauce/Vege 공통 셀 생성"""
        alignment = Qt.AlignLeft | Qt.AlignVCenter if col == 0 else Qt.AlignRight | Qt.AlignVCenter
        editable = col in (1, 4, 6)

        foreground = None
        if col == 5:
            try:
                if int(str(text).replace(",", "")) < 0:
                    foreground = QColor("#cc0000")
            except Exception:
                pass

        return self._create_cell(
            text=text, pk=pk, alignment=alignment,
            editable=editable, foreground=foreground,
        )

    def _apply_column_resize_rules(self):
        idx = self.ui.tabWidget.currentIndex()
        tables = {0: self.ui.tableWidget1, 1: self.ui.tableWidget2,
                  2: self.ui.tableWidget3, 3: self.ui.tableWidget4}
        table = tables.get(idx)
        if not table:
            return

        header = table.horizontalHeader()
        col_count = table.columnCount()

        table.resizeColumnsToContents()

        name_col = deadline_col = None
        for col in range(col_count):
            item = table.horizontalHeaderItem(col)
            if not item:
                continue
            text = item.text().strip()
            if text == "품명":
                name_col = col
            elif text == "소비기한":
                deadline_col = col

        for c in range(col_count):
            header.setSectionResizeMode(c, QHeaderView.Stretch)

        if name_col is not None:
            header.setSectionResizeMode(name_col, QHeaderView.Fixed)
            table.setColumnWidth(name_col, 540)

        if deadline_col is not None:
            header.setSectionResizeMode(deadline_col, QHeaderView.Fixed)
            table.setColumnWidth(deadline_col, 160)

        header.setMinimumSectionSize(10)

    def _apply_column_visibility_rules(self):
        table = self.ui.tableWidget1

        admin_only_cols = [
            COL_VENDOR, COL_PKG, COL_PREV_RES, COL_PRODUCTION,
            COL_PLAN_KG, COL_TODAY_RES
        ]

        # 업체명은 무조건 숨김
        table.setColumnHidden(COL_VENDOR, True)

        for col in admin_only_cols:
            if col == COL_VENDOR:
                continue
            table.setColumnHidden(col, self.current_level < 1)

    # =====================================================
    # 5. 데이터 로딩
    # =====================================================
    def _load_product_tab(self):
        table = self.ui.tableWidget1
        qdate: QDate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")
        self.ui.label_retailer.setText(self.current_vendor)

        try:
            with db_connection(DB_NAME) as (conn, cur):
                sql = """
                    SELECT
                        A.PK, A.co, A.rname, A.uname, A.pkg,
                        A.order_qty, A.order_qty_after,
                        A.prev_residue, A.production_plan,
                        A.produced_qty, A.today_residue,
                        A.work_status,
                        B.deadline,
                        A.recent_chulgo
                    FROM ORDER_DASHBOARD A
                    LEFT JOIN Dashboard_UNAME_MAP B
                           ON A.uname = B.before_value
                          AND A.rname = B.retailer
                    WHERE CONVERT(DATE, A.sdate) = %s
                """
                params = [sdate_str]

                if not self.show_hidden:
                    sql += " AND (A.hide = 0 OR A.hide IS NULL)"

                if self.current_vendor == "코스트코":
                    sql += " AND A.rname IN ('코스트코', '코스온')"
                elif self.current_vendor == "홈플/컬리":
                    sql += " AND A.rname IN ('홈플러스', '마켓컬리')"
                else:
                    sql += " AND A.rname = %s"
                    params.append(self.current_vendor)

                sql += " ORDER BY A.RNAME DESC, A.PK"
                df = runquery(cur, sql, params)
        except (ConnectionError, Exception) as e:
            print(f"[_load_product_tab] DB 연결 실패: {e}")
            table.blockSignals(True)
            self._setup_product_headers(table)
            table.setRowCount(0)
            table.blockSignals(False)
            return

        table.blockSignals(True)
        self._setup_product_headers(table)
        table.setRowCount(0)

        if df is None or len(df) == 0:
            self.product_total_pages = 1
            self.product_page = 0
            self._update_pagination_ui()
            table.blockSignals(False)
            return

        df = pd.DataFrame(df)
        df.columns = [str(c).upper() for c in df.columns]

        # 페이지네이션 적용
        total_rows = len(df)
        self.product_total_pages = max(1, ceil(total_rows / self.product_page_size))
        if self.product_page >= self.product_total_pages:
            self.product_page = self.product_total_pages - 1
        start = self.product_page * self.product_page_size
        end = min(start + self.product_page_size, total_rows)
        df_page = df.iloc[start:end]

        table.setRowCount(len(df_page))

        for row_idx, row in enumerate(df_page.itertuples(index=False)):
            pk = int(row.PK)
            co_val = str(row.CO).strip()

            rname = row.RNAME.strip() if row.RNAME else ""
            uname_raw = row.UNAME.strip() if row.UNAME else ""
            uname = self.uname_map_cache.get(uname_raw, uname_raw)

            pkg = float(row.PKG)
            order_qty = int(row.ORDER_QTY)
            order_qty_after = int(row.ORDER_QTY_AFTER)
            prev_residue = int(row.PREV_RESIDUE)
            produced_qty = int(row.PRODUCED_QTY)
            today_residue = int(row.TODAY_RESIDUE)
            production_plan = int(row.PRODUCTION_PLAN)

            # 최근출고 시각 포맷팅
            recent_chulgo_val = row.RECENT_CHULGO
            shipment_time_str = "-"
            if recent_chulgo_val:
                try:
                    s_val = str(recent_chulgo_val)
                    if len(s_val) >= 16:
                        shipment_time_str = s_val[11:16]
                except Exception:
                    pass

            # 계산 필드
            diff = order_qty_after - order_qty
            diff_display = "" if diff == 0 else str(diff)

            production_qty = max(order_qty_after - prev_residue, 0)
            plan_qty = production_plan
            plan_kg = plan_qty * pkg

            # 수율 계산 (cal_values에서 가져옴)
            trate_val = calc_trate_value(
                co=co_val,
                order_qty_after=order_qty_after,
                prev_residue=prev_residue,
                today_residue=today_residue,
                production_plan=production_plan,
                sdate_str=sdate_str,
            )
            if trate_val is None:
                trate_text = "-"
                trate_color = None
            else:
                trate_text = f"{trate_val:.1f}"
                trate_int = int(trate_val)
                trate_color = QColor("#cc0000") if (trate_int < 90 or trate_int >= 100) else None

            # 작업상태 자동 계산
            if plan_qty <= 0:
                work_status = "-"
            elif produced_qty > order_qty_after:
                work_status = "초과"
            elif produced_qty == order_qty_after:
                work_status = "완료"
            else:
                work_status = ""

            # 소비기한 계산
            deadline_val = ""
            if row.DEADLINE is not None and not pd.isna(row.DEADLINE):
                try:
                    days = int(float(row.DEADLINE))
                    calc_date = qdate.addDays(days - 1)
                    deadline_val = calc_date.toString("yy-MM-dd")
                except Exception:
                    deadline_val = ""

            values = [
                rname, uname, deadline_val, fmt(f"{pkg:.1f}"),
                fmt(order_qty), fmt(order_qty_after), fmt(diff_display),
                fmt(prev_residue), fmt(production_qty), fmt(plan_qty),
                fmt(round(plan_kg)), fmt(produced_qty), shipment_time_str,
                fmt(today_residue), trate_text, work_status,
            ]

            for col, text in enumerate(values):
                item = self._create_product_item(text, pk, col)
                item.setData(Qt.UserRole + 10, co_val)

                if col == COL_TRATE and trate_color:
                    item.setForeground(QBrush(trate_color))

                table.setItem(row_idx, col, item)

        self._apply_column_resize_rules()

        if not self._item_changed_connected["product"]:
            table.itemChanged.connect(self._on_product_item_changed)
            self._item_changed_connected["product"] = True

        table.blockSignals(False)
        self._apply_column_visibility_rules()

        # 최근출고(물류용) 모드
        is_logistics_mode = self.ui.ml_check.isChecked()
        table.setColumnHidden(COL_SHIPMENT_TIME, not is_logistics_mode)
        table.setColumnHidden(COL_TRATE, is_logistics_mode)

        # 페이지네이션 바 업데이트
        self._update_pagination_ui()

    def _load_material_tab(self, tab_key: str):
        """Raw/Sauce/Vege 탭 공통 로딩"""
        table = self._material_table(tab_key)
        db_table = self._material_db_table(tab_key)
        row_height = self._material_row_height(tab_key)
        sdate_str = self.ui.dateEdit.date().toString("yyyy-MM-dd")

        table.blockSignals(True)
        self._setup_material_headers(table)
        table.setRowCount(0)

        try:
            with db_connection(DB_NAME) as (conn, cur):
                sql = f"""
                    SELECT PK, uname, co, stock, order_qty,
                           order_qty_after, prepro_qty, ipgo_qty
                    FROM {db_table}
                    WHERE CONVERT(DATE, sdate) = %s
                    ORDER BY uname, co, PK
                """
                df = runquery(cur, sql, [sdate_str])
        except (ConnectionError, Exception) as e:
            print(f"[_load_material_tab] DB 연결 실패: {e}")
            table.blockSignals(False)
            return

        if df is None or len(df) == 0:
            table.blockSignals(False)
            return

        df = pd.DataFrame(df)
        df.columns = [str(c).upper() for c in df.columns]
        table.setRowCount(len(df))

        for row_idx, row in enumerate(df.itertuples(index=False)):
            pk = int(row.PK)
            stock = int(row.STOCK)
            order_qty = int(row.ORDER_QTY)
            order_qty_after = int(row.ORDER_QTY_AFTER)
            prepro_qty = int(row.PREPRO_QTY)
            ipgo_qty = int(row.IPGO_QTY)

            expected_short = stock - order_qty_after - prepro_qty
            expected_stock = expected_short + ipgo_qty

            row_values = [
                str(row.UNAME).strip(),
                fmt(stock), fmt(order_qty), fmt(order_qty_after),
                fmt(prepro_qty), fmt(expected_short),
                fmt(ipgo_qty), fmt(expected_stock),
            ]

            for col_idx, value in enumerate(row_values):
                item = self._create_material_item(value, pk, col_idx)
                table.setItem(row_idx, col_idx, item)

        table.verticalHeader().setDefaultSectionSize(row_height)
        self._apply_column_resize_rules()

        if not self._item_changed_connected[tab_key]:
            handler = lambda item, k=tab_key: self._on_material_item_changed(k, item)
            table.itemChanged.connect(handler)
            self._item_changed_connected[tab_key] = True

        table.blockSignals(False)

    # =====================================================
    # 단일 행 갱신
    # =====================================================
    def _refresh_single_row(self, pk: int):
        table = self.ui.tableWidget1
        qdate: QDate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")

        with db_connection(DB_NAME) as (conn, cur):
            sql = """
                SELECT
                    PK, co, rname, uname, pkg,
                    order_qty, order_qty_after,
                    prev_residue, production_plan, produced_qty,
                    today_residue, recent_chulgo
                FROM ORDER_DASHBOARD
                WHERE PK = %s
            """
            df = runquery(cur, sql, [pk])

        if df is None or len(df) == 0:
            return

        r = pd.DataFrame(df)
        r.columns = [str(c).upper() for c in r.columns]
        r = r.iloc[0]

        co_val = str(r.get("CO", "") or "").strip()
        order_qty = r["ORDER_QTY"]
        order_qty_after = r["ORDER_QTY_AFTER"]
        prev_residue = r["PREV_RESIDUE"]
        today_residue = r["TODAY_RESIDUE"]
        production_plan = r["PRODUCTION_PLAN"]
        produced_qty = r["PRODUCED_QTY"]
        pkg = r["PKG"]

        production_qty = max(order_qty_after - prev_residue, 0)
        plan_kg = production_plan * pkg
        diff = order_qty_after - order_qty

        trate_val = calc_trate_value(
            co=co_val,
            order_qty_after=order_qty_after,
            prev_residue=prev_residue,
            today_residue=today_residue,
            production_plan=production_plan,
            sdate_str=sdate_str,
        )
        trate_text = "-" if trate_val is None else f"{trate_val:.1f}"

        if production_plan <= 0:
            work_status = "-"
        elif produced_qty > order_qty_after:
            work_status = "초과"
        elif produced_qty == order_qty_after:
            work_status = "완료"
        else:
            work_status = ""

        recent_chulgo_val = r.get("RECENT_CHULGO")
        shipment_time_str = "-"
        if recent_chulgo_val:
            s_val = str(recent_chulgo_val)
            if len(s_val) >= 16:
                shipment_time_str = s_val[11:16]

        values = {
            COL_VENDOR: r["RNAME"],
            COL_PRODUCT: r["UNAME"],
            COL_PKG: fmt(f"{pkg:.1f}"),
            COL_ORDER: fmt(order_qty),
            COL_FINAL_ORDER: fmt(order_qty_after),
            COL_DIFF: "" if diff == 0 else fmt(diff),
            COL_PREV_RES: fmt(prev_residue),
            COL_PRODUCTION: fmt(production_qty),
            COL_PLAN: fmt(production_plan),
            COL_PLAN_KG: fmt(f"{plan_kg:.1f}"),
            COL_CUR_PROD: fmt(produced_qty),
            COL_SHIPMENT_TIME: shipment_time_str,
            COL_TODAY_RES: fmt(today_residue),
            COL_TRATE: trate_text,
            COL_WORK_STATUS: work_status,
        }

        row_idx = -1
        for i in range(table.rowCount()):
            if table.item(i, 0) and table.item(i, 0).data(Qt.UserRole) == pk:
                row_idx = i
                break

        if row_idx == -1:
            return

        table.blockSignals(True)
        for col, text in values.items():
            item = self._create_product_item(text, pk, col)
            table.setItem(row_idx, col, item)
        table.blockSignals(False)

    def _refresh_single_material_row(self, tab_key: str, pk: int):
        """Raw/Sauce/Vege 단일 행 갱신 공통"""
        table = self._material_table(tab_key)
        db_table = self._material_db_table(tab_key)

        with db_connection(DB_NAME) as (conn, cur):
            sql = f"""
                SELECT PK, uname, stock, order_qty, order_qty_after,
                       prepro_qty, ipgo_qty
                FROM {db_table}
                WHERE PK = %s
            """
            df = runquery(cur, sql, [pk])

        if df is None or df.empty:
            return

        r = pd.DataFrame(df)
        r.columns = [str(c).upper() for c in r.columns]
        r = r.iloc[0]

        stock = int(r["STOCK"])
        order_qty = int(r["ORDER_QTY"])
        order_qty_after = int(r["ORDER_QTY_AFTER"])
        prepro_qty = int(r["PREPRO_QTY"])
        ipgo_qty = int(r["IPGO_QTY"])

        expected_short = stock - order_qty_after - prepro_qty
        expected_stock = expected_short + ipgo_qty

        values = [
            r["UNAME"], fmt(stock), fmt(order_qty), fmt(order_qty_after),
            fmt(prepro_qty), fmt(expected_short), fmt(ipgo_qty), fmt(expected_stock),
        ]

        row_idx = -1
        for i in range(table.rowCount()):
            if table.item(i, 0) and table.item(i, 0).data(Qt.UserRole) == pk:
                row_idx = i
                break

        if row_idx == -1:
            return

        table.blockSignals(True)
        for col, v in enumerate(values):
            item = self._create_material_item(str(v), pk, col)
            table.setItem(row_idx, col, item)
        table.blockSignals(False)

    # =====================================================
    # 6. 테이블 수정 이벤트 처리
    # =====================================================
    def _on_product_item_changed(self, item: QTableWidgetItem):
        col = item.column()
        if col not in (COL_PLAN, COL_TODAY_RES, COL_PREV_RES):
            return

        pk = item.data(Qt.UserRole)
        if pk is None:
            return

        text = item.text().replace(",", "").strip()
        try:
            new_val = int(text) if text else 0
            if new_val < 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "오류", "0 이상 정수만 입력 가능합니다.")
            self.ui.tableWidget1.blockSignals(True)
            item.setText(fmt(0))
            self.ui.tableWidget1.blockSignals(False)
            new_val = 0

        field_map = {
            COL_PLAN: "production_plan",
            COL_TODAY_RES: "today_residue",
            COL_PREV_RES: "prev_residue",
        }
        field_name = field_map.get(col)
        if not field_name:
            return

        with db_connection(DB_NAME) as (conn, cur):
            # 변경 전 값 조회
            old_val = 0
            try:
                df_old = runquery(cur, f"SELECT {field_name} FROM ORDER_DASHBOARD WHERE PK = %s", [pk])
                if df_old is not None and not df_old.empty:
                    old_val = int(df_old.iloc[0, 0] or 0)
            except Exception:
                pass

            sql = f"UPDATE ORDER_DASHBOARD SET {field_name} = %s WHERE PK = %s"
            runquery(cur, sql, [new_val, pk])

            # 로그 기록
            if old_val != new_val:
                row = item.row()
                u_item = self.ui.tableWidget1.item(row, COL_PRODUCT)
                uname = u_item.text() if u_item else "-"

                label_map = {
                    "production_plan": "생산계획",
                    "today_residue": "당일잔피",
                    "prev_residue": "전일잔피",
                }
                lbl = label_map.get(field_name, field_name)
                content = f"{lbl} {old_val} -> {new_val}"
                DashboardLogDialog.log_change(
                    self.current_user, self.ui.dateEdit.date(), uname, content, ""
                )

        self._refresh_single_row(pk)

    def _on_material_item_changed(self, tab_key: str, item: QTableWidgetItem):
        """Raw/Sauce/Vege 아이템 변경 공통 핸들러"""
        col = item.column()
        if col not in (1, 4, 6):
            return

        table = self._material_table(tab_key)
        db_table = self._material_db_table(tab_key)
        row = item.row()
        pk = item.data(Qt.UserRole)
        if pk is None:
            return

        def get_int(c):
            v = table.item(row, c)
            if not v:
                return 0
            try:
                return int(str(v.text()).replace(",", ""))
            except Exception:
                return 0

        stock = get_int(1)
        prepro = get_int(4)
        incoming = get_int(6)

        with db_connection(DB_NAME) as (conn, cur):
            # 로그용: 변경 전 값 조회
            old_vals = {}
            try:
                df_old = runquery(
                    cur, f"SELECT stock, prepro_qty, ipgo_qty FROM {db_table} WHERE PK = %s", [pk]
                )
                if df_old is not None and not df_old.empty:
                    old_vals["stock"] = int(df_old.iloc[0][0] or 0)
                    old_vals["prepro_qty"] = int(df_old.iloc[0][1] or 0)
                    old_vals["ipgo_qty"] = int(df_old.iloc[0][2] or 0)
            except Exception:
                pass

            sql = f"""
                UPDATE {db_table}
                SET stock = %s, prepro_qty = %s, ipgo_qty = %s
                WHERE PK = %s
            """
            runquery(cur, sql, [stock, prepro, incoming, pk])

            # 로그 기록
            changed_content = []
            if old_vals.get("stock", -999) != stock:
                changed_content.append(f"재고 {old_vals.get('stock')}->{stock}")
            if old_vals.get("prepro_qty", -999) != prepro:
                changed_content.append(f"선생산 {old_vals.get('prepro_qty')}->{prepro}")
            if old_vals.get("ipgo_qty", -999) != incoming:
                changed_content.append(f"입고 {old_vals.get('ipgo_qty')}->{incoming}")

            if changed_content:
                u_item = table.item(row, 0)
                uname = u_item.text() if u_item else "-"
                content = ", ".join(changed_content)
                DashboardLogDialog.log_change(
                    self.current_user, self.ui.dateEdit.date(), uname, content, ""
                )

        self._refresh_single_material_row(tab_key, pk)

    # =====================================================
    # 8. 대시보드 데이터 가공
    # =====================================================
    def _generate_material_dashboard(self, db_table: str, recipe_keyword: str, bco_list: list):
        """Raw/Sauce 대시보드 공통 생성 로직"""
        qdate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")
        sdate_dt = datetime(qdate.year(), qdate.month(), qdate.day(), 0, 0, 0)
        now = datetime.now()

        with db_connection(DB_NAME) as (conn, cur):
            sql_order = """
                SELECT co, order_qty_after, production_plan, prev_residue, pkg
                FROM ORDER_DASHBOARD
                WHERE CONVERT(DATE, sdate) = %s
            """
            df_order = runquery(cur, sql_order, [sdate_str])

        if df_order is None or df_order.empty:
            with db_connection(DB_NAME) as (conn, cur):
                runquery(cur, f"DELETE FROM {db_table} WHERE CONVERT(DATE, sdate) = %s", [sdate_str])
            return

        df_order.columns = [c.upper() for c in df_order.columns]
        df_order["CO"] = df_order["CO"].astype(str).str.strip()

        grouped = calc_plan_kg_by_recipe(df_order, recipe_keyword, bco_list)

        if grouped is None or grouped.empty:
            with db_connection(DB_NAME) as (conn, cur):
                runquery(cur, f"DELETE FROM {db_table} WHERE CONVERT(DATE, sdate) = %s", [sdate_str])
            return

        with db_connection(DB_NAME) as (conn, cur):
            runquery(cur, f"DELETE FROM {db_table} WHERE CONVERT(DATE, sdate) = %s", [sdate_str])

        rows = []
        for _, r in grouped.iterrows():
            bco = str(r["BCO"]).strip()
            buname = str(r["BUNAME"]).strip()
            plan_kg_sum = float(r["PLAN_KG"] or 0)
            qty_int = int(round(plan_kg_sum))

            if qty_int <= 0:
                continue

            stock_val = get_stock_from_pan(bco, sdate_str)
            rows.append({
                "uname": buname, "co": bco, "sdate": sdate_dt,
                "created_time": now, "stock": stock_val,
                "order_qty": qty_int, "order_qty_after": qty_int,
                "prepro_qty": 0, "ipgo_qty": 0,
            })

        if rows:
            self._insert_material_rows(db_table, rows)

    def _dashboard_vege_from_dashboard(self):
        qdate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")
        sdate_dt = datetime(qdate.year(), qdate.month(), qdate.day(), 0, 0, 0)
        now = datetime.now()

        VEGE_BCO_LIST = ["720192", "700122", "720094", "710665"]

        with db_connection(DB_NAME) as (conn, cur):
            sql = """
                SELECT co, order_qty_after, production_plan, prev_residue, pkg
                FROM ORDER_DASHBOARD
                WHERE CONVERT(DATE, sdate) = %s
            """
            df_order = runquery(cur, sql, [sdate_str])

        if df_order is None or df_order.empty:
            with db_connection(DB_NAME) as (conn, cur):
                runquery(cur, "DELETE FROM DASHBOARD_VEGE WHERE CONVERT(DATE, sdate) = %s", [sdate_str])
            return

        df_order.columns = [c.upper() for c in df_order.columns]
        df_order["CO"] = df_order["CO"].astype(str).str.strip()

        co_list = df_order["CO"].unique().tolist()
        if not co_list:
            return

        with db_connection("GFOOD_B") as (conn, cur):
            sql = f"""
                SELECT CO, BCO, BUNAME, SA
                FROM RECIPE
                WHERE BCO IN ({','.join(['%s'] * len(VEGE_BCO_LIST))})
                  AND CO IN ({','.join(['%s'] * len(co_list))})
            """
            df_recipe = runquery(cur, sql, VEGE_BCO_LIST + co_list)

        if df_recipe is None or df_recipe.empty:
            with db_connection(DB_NAME) as (conn, cur):
                runquery(cur, "DELETE FROM DASHBOARD_VEGE WHERE CONVERT(DATE, sdate) = %s", [sdate_str])
            return

        df_recipe.columns = [c.upper() for c in df_recipe.columns]
        df_recipe["CO"] = df_recipe["CO"].astype(str).str.strip()
        df_recipe["BCO"] = df_recipe["BCO"].astype(str).str.strip()

        df = df_order.merge(df_recipe, on="CO", how="inner")
        if df.empty:
            return

        df["PLAN_KG"] = df["PRODUCTION_PLAN"].fillna(0).astype(float) * df["PKG"].fillna(0).astype(float)
        df = df[df["PLAN_KG"] > 0]
        if df.empty:
            return

        df["VEGE_KG"] = df["PLAN_KG"] * df["SA"].fillna(0).astype(float)
        df = df[df["VEGE_KG"] > 0]
        if df.empty:
            return

        grouped = df.groupby(["BCO", "BUNAME"], as_index=False)["VEGE_KG"].sum()

        with db_connection(DB_NAME) as (conn, cur):
            runquery(cur, "DELETE FROM DASHBOARD_VEGE WHERE CONVERT(DATE, sdate) = %s", [sdate_str])

        rows = []
        for _, r in grouped.iterrows():
            qty_int = int(round(float(r["VEGE_KG"] or 0)))
            if qty_int <= 0:
                continue

            stock_val = get_stock_from_pan(str(r["BCO"]), sdate_str)
            rows.append({
                "uname": r["BUNAME"], "co": r["BCO"], "sdate": sdate_dt,
                "created_time": now, "stock": stock_val,
                "order_qty": qty_int, "order_qty_after": qty_int,
                "prepro_qty": 0, "ipgo_qty": 0,
            })

        if rows:
            self._insert_material_rows("DASHBOARD_VEGE", rows)

    # =====================================================
    # 9. DB Insert/Update/Delete
    # =====================================================
    def _insert_dashboard_rows(self, rows):
        with db_connection(DB_NAME) as (conn, cur):
            sql = """
                INSERT INTO ORDER_DASHBOARD (
                    bigo, sdate, created_time, id,
                    rname, uname, co, pkg,
                    order_qty, order_qty_after, prev_residue, production_plan,
                    produced_qty, today_residue
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            for r in rows:
                params = [
                    r["bigo"], r["sdate"], r["created_time"], r["id"],
                    r["rname"], r["uname"], r["co"], r["pkg"],
                    r["order_qty"], r["order_qty_after"], r["prev_residue"],
                    r["production_plan"], r["produced_qty"], r["today_residue"],
                ]
                runquery(cur, sql, params)

    def _insert_material_rows(self, db_table: str, rows: list):
        """Raw/Sauce/Vege INSERT 공통"""
        with db_connection(DB_NAME) as (conn, cur):
            sql = f"""
                INSERT INTO {db_table} (
                    uname, co, sdate, created_time,
                    stock, order_qty, order_qty_after,
                    prepro_qty, ipgo_qty
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            for r in rows:
                runquery(cur, sql, [
                    r["uname"], r["co"], r["sdate"], r["created_time"],
                    r["stock"], r["order_qty"], r["order_qty_after"],
                    r["prepro_qty"], r["ipgo_qty"],
                ])

    def on_click_add_dummy_rows(self):
        dlg = ProductListDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return

        self.product_list = dlg.get_product_list()

        if not self.product_list:
            QMessageBox.information(self, "안내", "PRODUCT_LIST가 비어 있습니다.")
            return

        qdate: QDate = self.ui.dateEdit.date()
        sdate_dt = datetime(qdate.year(), qdate.month(), qdate.day(), 0, 0, 0)
        sdate_str = qdate.toString("yyyy-MM-dd")
        now = datetime.now()

        rows = []

        with db_connection("GFOOD_B") as (conn_master, cur_master):
            for base_co, vendor in self.product_list:
                base_co = str(base_co).strip()
                if not base_co:
                    continue

                df_master = runquery(
                    cur_master,
                    "SELECT TOP 1 CO, UNAME, PACKG, PACSU FROM MASTER WHERE CO = %s",
                    [base_co],
                )

                if df_master is None or df_master.empty:
                    print(f"[SKIP:MASTER NOT FOUND] vendor={vendor}  base_co={base_co}")
                    continue

                m = df_master.iloc[0]
                uname = str(m.get("UNAME", "")).strip()

                packg_raw = m.get("PACKG", None)
                pkg = 0.0
                if packg_raw is not None:
                    try:
                        pkg = float(packg_raw)
                    except Exception:
                        try:
                            pkg = float(str(packg_raw).replace("KG", "").replace("kg", "").strip())
                        except Exception:
                            pkg = 0.0

                pacsu_raw = m.get("PACSU", 1)
                try:
                    pacsu = int(pacsu_raw if pacsu_raw not in (None, "") else 1)
                except Exception:
                    pacsu = 1
                if pacsu <= 0:
                    pacsu = 1

                prev_residue = get_prev_residue_from_today(base_co)

                order_qty_packs = calc_order_qty_packs(
                    base_co=base_co, vendor=vendor,
                    sdate_str=sdate_str, pacsu=pacsu,
                )

                produced_qty_val, produced_time = get_produced_qty_packs(base_co, sdate_str, pacsu)

                rows.append({
                    "bigo": "", "sdate": sdate_dt, "created_time": now,
                    "id": "인길환", "rname": vendor, "uname": uname,
                    "co": base_co, "pkg": pkg,
                    "order_qty": order_qty_packs, "order_qty_after": order_qty_packs,
                    "prev_residue": prev_residue, "production_plan": 0,
                    "produced_qty": produced_qty_val, "today_residue": 0,
                })

        if not rows:
            QMessageBox.information(self, "안내", "INSERT할 데이터가 없습니다.")
            return

        try:
            self._insert_dashboard_rows(rows)
            self._generate_material_dashboard("DASHBOARD_RAW", "(정선)", ['502811'])
            self._generate_material_dashboard("DASHBOARD_SAUCE", "소스", ['600901'])
            self._dashboard_vege_from_dashboard()

            QMessageBox.information(
                self, "완료",
                f"제품 {len(rows)}행, 원료/소스/야채 대시보드 재생성 완료."
            )
            DashboardLogDialog.log_action(
                self.current_user, self.ui.dateEdit.date(),
                f"표 생성(dummy rows) {len(rows)}행"
            )
            if hasattr(self.ui, "tabWidget") and self.ui.tabWidget.currentIndex() == 0:
                self._load_product_tab()

        except Exception as e:
            QMessageBox.critical(self, "에러", str(e))

    def on_click_show_log_dialog(self):
        dlg = DashboardLogDialog(self)
        dlg.exec_()

    def on_click_delete_selected_products(self):
        """선택한 제품만 삭제 + RAW/SAUCE/VEGE 재집계"""
        table = self.ui.tableWidget1
        selected_rows = sorted({idx.row() for idx in table.selectedIndexes()})

        if not selected_rows:
            QMessageBox.information(self, "안내", "삭제할 제품을 선택하세요.")
            return

        UNAME_COL = 1
        uname_after_list = []
        for r in selected_rows:
            item = table.item(r, UNAME_COL)
            if item:
                uname_after_list.append(item.text().strip())

        if not uname_after_list:
            QMessageBox.warning(self, "오류", "선택한 행에서 제품명(UNAME)을 찾을 수 없습니다.")
            return

        uname_after_list = list(set(uname_after_list))

        # Dashboard_UNAME_MAP 조회하여 after → before 매핑
        with db_connection(DB_NAME) as (conn, cur):
            sql = "SELECT before_value, after_value FROM Dashboard_UNAME_MAP"
            df_map = runquery(cur, sql)

        mapping = {}
        if df_map is not None and not df_map.empty:
            for _, row in df_map.iterrows():
                bf = str(row["before_value"]).strip()
                af = str(row["after_value"]).strip()
                mapping[af] = bf

        uname_final_list = list(set(
            mapping.get(af, af) for af in uname_after_list
        ))

        reply = QMessageBox.question(
            self, "삭제 확인",
            f"선택한 {len(uname_after_list)}개의 제품을 삭제하시겠습니까?\n"
            f"(ORDER_DASHBOARD 삭제 + RAW/SAUCE/VEGE 재집계)",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        qdate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")

        with db_connection(DB_NAME) as (conn, cur):
            placeholders = ", ".join(["%s"] * len(uname_final_list))
            sql = f"""
                DELETE FROM ORDER_DASHBOARD
                WHERE CONVERT(DATE, sdate) = %s
                  AND UNAME IN ({placeholders})
            """
            runquery(cur, sql, [sdate_str] + uname_final_list)

        try:
            recalc_dashboard_raw_keep_manual(sdate_str)
            recalc_dashboard_sauce_keep_manual(sdate_str)
            recalc_dashboard_vege_keep_manual(sdate_str)
        except Exception as e:
            QMessageBox.critical(self, "재집계 오류", str(e))
            return

        QMessageBox.information(self, "완료", "선택한 제품이 삭제되었으며 재집계가 완료되었습니다.")
        DashboardLogDialog.log_action(
            self.current_user, self.ui.dateEdit.date(),
            f"선택 행 삭제 ({len(uname_final_list)}건)"
        )
        self._load_product_tab()

    def on_click_delete_rows(self):
        qdate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")

        reply = QMessageBox.question(
            self, "삭제 확인",
            f"{sdate_str} 데이터 전체를 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        with db_connection(DB_NAME) as (conn, cur):
            for tbl in ["ORDER_DASHBOARD", "DASHBOARD_RAW", "DASHBOARD_SAUCE", "DASHBOARD_VEGE"]:
                runquery(cur, f"DELETE FROM {tbl} WHERE CONVERT(DATE, sdate) = %s", [sdate_str])

        QMessageBox.information(self, "완료", f"{sdate_str} 자료 삭제 완료!")
        DashboardLogDialog.log_action(self.current_user, qdate, f"표 삭제 ({sdate_str})")

        for table in self._all_tables():
            table.setRowCount(0)

    # =====================================================
    # 생산량(produced_qty) 재계산 & UPDATE
    # =====================================================
    def on_click_update_product(self, checked=False, *, silent=False):
        try:
            qdate: QDate = self.ui.dateEdit.date()
            sdate_str = qdate.toString("yyyy-MM-dd")

            try:
                conn, cur = getdb(DB_NAME)
            except Exception as e:
                msg = f"{DB_NAME} 연결 실패:\n{e}"
                if not silent:
                    QMessageBox.critical(self, "DB 오류", msg)
                else:
                    print(f"[ERROR] {msg}")
                return

            try:
                sql = "SELECT DISTINCT co FROM ORDER_DASHBOARD WHERE CONVERT(DATE, sdate) = %s"
                df = runquery(cur, sql, [sdate_str])
            finally:
                closedb(conn)

            if df is None or len(df) == 0:
                if not silent:
                    QMessageBox.information(self, "안내", f"{sdate_str} 기준 데이터가 없습니다.")
                return

            df = pd.DataFrame(df)
            co_col = df.columns[0]

            try:
                conn_u, cur_u = getdb(DB_NAME)
            except Exception as e:
                msg = f"{DB_NAME} 연결 실패(UPDATE):\n{e}"
                if not silent:
                    QMessageBox.critical(self, "DB 오류", msg)
                return

            updated_cnt = 0
            try:
                for co_val in df[co_col]:
                    co_str = str(co_val).strip()
                    if not co_str:
                        continue

                    try:
                        pacsu = get_pacsu_by_co(co_str)
                    except Exception as e:
                        print(f"[ERROR] get_pacsu_by_co({co_str}) 예외: {e}")
                        pacsu = 1

                    produced_qty, recent_time_val = get_produced_qty_packs(co_str, sdate_str, pacsu)

                    try:
                        runquery(
                            cur_u,
                            """
                            UPDATE ORDER_DASHBOARD
                            SET produced_qty = %s, recent_chulgo = %s
                            WHERE CONVERT(DATE, sdate) = %s AND co = %s
                            """,
                            [produced_qty, recent_time_val, sdate_str, co_str],
                        )
                        updated_cnt += 1
                    except Exception as e:
                        print(f"[ERROR] produced_qty UPDATE 실패 co={co_str}: {e}")
                        continue
            finally:
                closedb(conn_u)

            msg = f"{sdate_str} 기준 {updated_cnt}개 품목의 생산 팩수(produced_qty)를 갱신했습니다."
            if not silent:
                QMessageBox.information(self, "완료", msg)
            else:
                print(f"[INFO] {msg}")
            self._load_product_tab()

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "예외 발생", f"생산량 갱신 중 예외가 발생했습니다.\n{e}")

    # =====================================================
    # 발주량 재계산 & UPDATE
    # =====================================================
    def on_click_update_order_qty_after(self, checked=False, *, silent=False):
        qdate: QDate = self.ui.dateEdit.date()
        sdate_str = qdate.toString("yyyy-MM-dd")

        if not self.product_list:
            if not silent:
                QMessageBox.information(self, "안내", "PRODUCT_LIST가 비어 있습니다.")
            return

        with db_connection(DB_NAME) as (conn, cur):
            for base_co, vendor in self.product_list:
                base_co = str(base_co).strip()

                pacsu = get_pacsu_by_co(base_co)
                if pacsu is None or pacsu <= 0:
                    pacsu = 1

                new_qty_packs = int(
                    calc_order_qty_packs(
                        base_co=base_co, vendor=vendor,
                        sdate_str=sdate_str, pacsu=pacsu,
                    )
                )

                runquery(cur, """
                    UPDATE ORDER_DASHBOARD
                    SET order_qty_after = %s
                    WHERE CONVERT(DATE, sdate) = %s AND co = %s
                """, [new_qty_packs, sdate_str, base_co])

        recalc_dashboard_raw_keep_manual(sdate_str)
        recalc_dashboard_sauce_keep_manual(sdate_str)
        recalc_dashboard_vege_keep_manual(sdate_str)

        msg = ("모든 제품의 최종 발주량(order_qty_after)이 재계산되었고,\n"
               "원료/소스/야채 대시보드도 최신 기준으로 반영되었습니다.")
        if not silent:
            QMessageBox.information(self, "완료", msg)
        else:
            print(f"[INFO] {msg.replace(chr(10), ' ')}")

        self._load_product_tab()

    def on_click_complete_product(self):
        table = self.ui.tableWidget1

        selected_rows = sorted({idx.row() for idx in table.selectedIndexes()})
        if not selected_rows:
            QMessageBox.information(self, "안내", "완료 처리할 제품 행을 선택하세요.")
            return

        with db_connection(DB_NAME) as (conn, cur):
            for row in selected_rows:
                item = table.item(row, 0)
                if not item:
                    continue

                pk = item.data(Qt.UserRole)
                if not pk:
                    continue

                runquery(
                    cur,
                    "UPDATE ORDER_DASHBOARD SET work_status = '완료' WHERE PK = %s",
                    [pk],
                )

                item_ws = table.item(row, COL_WORK_STATUS)
                if item_ws:
                    table.blockSignals(True)
                    item_ws.setText("완료")
                    table.blockSignals(False)

                self._refresh_single_row(pk)

        QMessageBox.information(self, "완료", "선택된 제품의 작업 상태가 '완료'로 변경되었습니다.")

    # =====================================================
    # 엑셀 내보내기
    # =====================================================
    def on_click_export_excel(self):
        import os
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Border, Side

        vendor_buttons = {
            "코스트코": self.on_click_filter_costco,
            "이마트": self.on_click_filter_emart,
            "홈플/컬리": self.on_click_filter_hk,
            "롯데": self.on_click_filter_lotte,
        }
        vendors = list(vendor_buttons.keys())

        today_str = datetime.now().strftime("%Y%m%d_%H%M")
        default_name = f"제품현황_업체별_{today_str}.xlsx"

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        default_path = os.path.join(desktop, default_name)

        path, _ = QFileDialog.getSaveFileName(
            self, "업체별 제품현황 엑셀 저장", default_path,
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:

                for vendor in vendors:
                    vendor_buttons[vendor]()
                    table = self.ui.tableWidget1

                    if table.rowCount() == 0:
                        continue

                    headers = []
                    for c in range(table.columnCount()):
                        item = table.horizontalHeaderItem(c)
                        headers.append(item.text() if item else f"열{c + 1}")

                    rows = []
                    for r in range(table.rowCount()):
                        row_vals = []
                        for c in range(table.columnCount()):
                            item = table.item(r, c)
                            row_vals.append(item.text() if item else "")
                        rows.append(row_vals)

                    df = pd.DataFrame(rows, columns=headers)
                    safe_name = vendor.replace("/", "_")
                    df.to_excel(writer, sheet_name=safe_name, index=False)

                    wb = writer.book
                    ws = wb[safe_name]

                    header_font = Font(bold=True)
                    header_align = Alignment(horizontal="center", vertical="center")
                    left_align = Alignment(horizontal="left", vertical="center")
                    right_align = Alignment(horizontal="right", vertical="center")
                    thin = Side(border_style="thin", color="000000")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    for col_idx, col_name in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.border = border

                        max_len = len(str(col_name))
                        col_series = df[col_name].astype(str)
                        if not col_series.empty:
                            max_len = max(max_len, col_series.map(len).max())
                        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

                    for row_idx in range(2, ws.max_row + 1):
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.border = border
                            if col_idx in (1, 2):
                                cell.alignment = left_align
                            else:
                                cell.alignment = right_align

            QMessageBox.information(self, "완료", f"엑셀 파일이 저장되었습니다.\n{path}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "오류", f"엑셀 저장 중 오류 발생\n{e}")


# ---------------------------------------------------------
# 실행
# ---------------------------------------------------------
if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        check_version_and_update(PROGRAM_NAME, CURRENT_VERSION)

        w = OrderDashboardWidget()

        screen = app.primaryScreen().availableGeometry()
        w.resize(screen.width(), screen.height())

        w.show()
        sys.exit(app.exec_())

    except Exception:
        import traceback
        print("\n===== 실행 중 오류 발생 =====")
        print(traceback.format_exc())
        input("\n 엔터를 누르면 닫힙니다...")
